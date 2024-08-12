
# careting folder taking name from excel file.

import shutil
import os
from openpyxl import load_workbook

def create_folders_from_excel(file_path, sheet_name, column_number):
    # Load the workbook
    wb = load_workbook(filename=file_path)
    
    # Select the active sheet
    sheet = wb[sheet_name]
    
    # Initialize list to store folder names
    folder_names = []
    
    
    # Read folder names from the specified column
    for row in sheet.iter_rows(min_row=2, min_col=column_number, max_col=column_number, values_only=True):
        folder_name = row[0]
        if folder_name:
            folder_names.append(folder_name.strip())  # Strip whitespace and add to list
    print(len(folder_names)) 
  
    # Create folders
    folder_path= input("Enter the path where folder need to be create: ")
    #base_dir = os.path.dirname(folder_path)
    for folder_name in folder_names:
        folder_destination_path = os.path.join(folder_path, folder_name)
        try:
            os.makedirs(folder_destination_path)
            
            #print(f"Folder '{folder_name}' created successfully at '{destination_path}'")
        except FileExistsError:
            print(f"Folder '{folder_name}' already exists at '{folder_destination_path}'")
        except Exception as e:
            print(f"Failed to create folder '{folder_name}' at '{folder_destination_path}': {str(e)}")
        # Define the directory path
    directory_path = input("Enter the folder path having files: ") #"C:\\Users\\V1\\OneDrive - Capgemini\\Desktop\\FSD & Design\\Mediasaturn"

    # Get the list of all entries in the directory
    all_entries = os.listdir(directory_path)

    # Filter out directories and get only file names
    n=0
    file_names = [entry for entry in all_entries if os.path.isfile(os.path.join(directory_path, entry))]
    for folder_name in folder_names:
        for file in file_names:
            if folder_name in file:
                #print("{} : :{}".format(folder_name,file))
                source= os.path.join(directory_path,file)
                destination= os.path.join(folder_path,folder_name)
                shutil.copy(source, destination)
                print("Souce & Destination: {} ; {}".format(source,destination))
                n=n+1
    print(n)

 
    

    # Print file names
    #for file_name in file_names:
        #print(file_name)

# Example usage:
if __name__ == "__main__":
    def user_value():
        excel_file=input("Enter excel file path having folder names: ")  #"C:\Users\V1\Downloads\FDT & Mediasaturn enhancements.xlsx"
        sheet_name= input("Please enter the sheet name: ")
        column_number=int(input("Please emter the coulumn number: "))
        create_folders_from_excel(excel_file, sheet_name, column_number)
    #excel_file = "C:\\Users\\V1\\Downloads\\FDT & Mediasaturn enhancements.xlsx"
    #sheet_name = "Page1"  # Replace with your sheet name
    #column_number = 1  # Column number where folder names are listed (A=1, B=2, etc.)
    user_value()
    
